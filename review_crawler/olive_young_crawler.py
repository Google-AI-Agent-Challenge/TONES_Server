import time
import re
import sys
import os
import json
import random
import pandas as pd
try:
    from curl_cffi import requests  # noqa: F401 (reserved for future use)
except ImportError:
    requests = None
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By

sys.stdout.reconfigure(encoding='utf-8')

# ==============================================================================
# CONFIGURATION
# ==============================================================================
OUTPUT_FILENAME = os.path.join(os.path.dirname(__file__), "스킨푸드_패드_고객리뷰.xlsx")
EDGE_BINARY_PATH = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
# ==============================================================================

TARGET_PADS = {
    "아스파라거스 패드": {"keywords": ["아스파라거스", "asparagus"],                     "default_goods": "A000000166709"},
    "복숭아 패드":       {"keywords": ["복숭아", "피치", "peach"],                       "default_goods": "A000000231714"},
    "블루 캐모마일 패드": {"keywords": ["캐모마일", "chamomile", "카모마일", "블루캐모마일"], "default_goods": "A000000166709"},
    "라이스 패드":       {"keywords": ["라이스", "rice", "쌀"],                          "default_goods": "A000000166709"},
    "레몬그라스 패드":   {"keywords": ["레몬그라스", "lemongrass"],                      "default_goods": "A000000166709"},
    "샤인머스캣 패드":   {"keywords": ["샤인머스캣", "shine", "머스캣"],                 "default_goods": "A000000166709"},
    "핑크자몽 패드":     {"keywords": ["자몽", "grapefruit", "핑크자몽"],                "default_goods": "A000000166709"},
    "미나리 패드":       {"keywords": ["미나리", "파슬리", "parsley", "판토텐산"],        "default_goods": "A000000185135"},
    "당근 패드":         {"keywords": ["당근", "캐롯", "carrot"],                        "default_goods": "A000000248098"},
    "감자 패드":         {"keywords": ["감자", "포테이토", "potato"],                    "default_goods": "A000000200396"},
    "도토리 패드":       {"keywords": ["도토리", "에이콘", "acorn"],                     "default_goods": "A000000157075"},
}

PRODUCT_PAGES = {
    "A000000166709": {"name": "11종 통합 기획전 페이지", "max_scroll_steps": 600, "target_reviews": 1500, "delay": 1.0},
    "A000000206889": {"name": "스킨푸드 패드 레시피 3종 페이지", "max_scroll_steps": 300, "target_reviews": 500, "delay": 1.0},
    "A000000231714": {"name": "복숭아 패드 전용 페이지",  "max_scroll_steps": 200, "target_reviews": 250,  "delay": 0.9},
    "A000000185135": {"name": "미나리 패드 전용 페이지",  "max_scroll_steps": 200, "target_reviews": 250,  "delay": 0.9},
    "A000000248098": {"name": "당근 패드 기획전 페이지",  "max_scroll_steps": 200, "target_reviews": 250,  "delay": 0.9},
    "A000000200396": {"name": "감자 패드 전용 페이지",    "max_scroll_steps": 200, "target_reviews": 250,  "delay": 0.9},
    "A000000157075": {"name": "도토리 패드 전용 페이지",  "max_scroll_steps": 200, "target_reviews": 250,  "delay": 0.9},
}

# ==============================================================================
# JavaScript: XHR + fetch 인터셉터
# review 탭 클릭 전에 주입 → 이후 모든 /review/api/ 요청 본문을 캡처
# ==============================================================================
_INTERCEPTOR_JS = """
if (!window.__oyIntercepted) {
    window.__oyIntercepted = true;
    window.__oyBatches    = [];

    // ── XHR ─────────────────────────────────────────────────────────
    const _xOpen = XMLHttpRequest.prototype.open;
    const _xSend = XMLHttpRequest.prototype.send;
    XMLHttpRequest.prototype.open = function(m, u) {
        this._oy_url = String(u || '');
        return _xOpen.apply(this, arguments);
    };
    XMLHttpRequest.prototype.send = function() {
        this.addEventListener('load', function() {
            if ((this._oy_url || '').includes('/review/api/')) {
                window.__oyBatches.push({
                    src: 'xhr', url: this._oy_url,
                    status: this.status, body: this.responseText || ''
                });
            }
        });
        return _xSend.apply(this, arguments);
    };

    // ── fetch ────────────────────────────────────────────────────────
    const _origFetch = window.fetch;
    window.fetch = async function(input, init) {
        const url = typeof input === 'string' ? input : ((input && input.url) || '');
        const res = await _origFetch.apply(this, arguments);
        if (url.includes('/review/api/')) {
            try {
                const text = await res.clone().text();
                window.__oyBatches.push({
                    src: 'fetch', url: url,
                    status: res.status, body: text || ''
                });
            } catch(e) {}
        }
        return res;
    };
}
"""

# 캡처된 배치를 수집하고 버퍼를 비움
_COLLECT_BATCHES_JS = """
const b = (window.__oyBatches || []).slice();
window.__oyBatches = [];
return JSON.stringify(b);
"""

# ==============================================================================
# JavaScript: Shadow DOM 직접 추출
# 렌더링된 oy-review-review-item 요소에서 데이터 파싱
# ==============================================================================
_EXTRACT_JS = """
(function(goodsNo) {
    const reviews = [];
    const seen    = new Set();

    function findAll(root, tagLower, depth) {
        if (!root || depth > 9) return [];
        let found = [];
        try {
            const all = Array.from(root.querySelectorAll ? root.querySelectorAll('*') : []);
            for (const el of all) {
                if (el.tagName && el.tagName.toLowerCase() === tagLower) found.push(el);
                if (el.shadowRoot) found = found.concat(findAll(el.shadowRoot, tagLower, depth + 1));
            }
        } catch(e) {}
        return found;
    }

    const items = findAll(document, 'oy-review-review-item', 0);

    for (const item of items) {
        if (!item.shadowRoot) continue;
        const sr = item.shadowRoot;

        const dateEl = sr.querySelector('span.date')
                    || sr.querySelector('.date')
                    || sr.querySelector('[class*="date"]')
                    || sr.querySelector('[class*="Date"]');
        const date = dateEl ? dateEl.textContent.trim() : '';

        const optEl = sr.querySelector('div.goods-option');
        let option  = optEl ? optEl.textContent.trim() : '단품';
        option = option.replace(/^\\[옵션\\]\\s*/, '').trim() || '단품';

        const ratingDiv = sr.querySelector('div.rating');
        let rating = 5;
        if (ratingDiv) {
            const stars = Array.from(ratingDiv.querySelectorAll('oy-review-star-icon'));
            const filledCount = stars.filter(s => {
                if (s.hasAttribute('empty'))            return false;
                if (s.getAttribute('type') === 'empty') return false;
                if (s.shadowRoot) {
                    const paths = Array.from(s.shadowRoot.querySelectorAll('path[fill], circle[fill]'));
                    for (const p of paths) {
                        const f = (p.getAttribute('fill') || '').toLowerCase();
                        if (f === '#d9d9d9' || f === 'gray' || f === '#ccc' || f === '#cccccc') return false;
                    }
                }
                return true;
            }).length;
            rating = filledCount > 0 ? filledCount : stars.length;
            if (rating < 1 || rating > 5) rating = 5;
        }

        let content = '';
        const contentComp = sr.querySelector('oy-review-review-content');
        if (contentComp && contentComp.shadowRoot) {
            const csr = contentComp.shadowRoot;
            const cEl = csr.querySelector('.review-content-container')
                     || csr.querySelector('[class*="content-container"]')
                     || csr.querySelector('[class*="review-content"]');
            content = cEl ? cEl.textContent.trim() : csr.textContent.trim();
            content = content.replace(/해당 리뷰는 성분과 내용물이 동일한[\\s\\S]*?있습니다\\./, '').trim();
        }
        if (!content || content.length < 5) continue;

        let username = '익명';
        let skinTypes = '';
        const userComp = sr.querySelector('oy-review-review-user');
        if (userComp && userComp.shadowRoot) {
            const uEl = userComp.shadowRoot.querySelector('.name')
                     || userComp.shadowRoot.querySelector('[class*="nickname"]')
                     || userComp.shadowRoot.querySelector('[class*="user-id"]')
                     || userComp.shadowRoot.querySelector('[class*="author"]');
            if (uEl) username = uEl.textContent.trim() || '익명';

            // 피부타입 추출 보강
            const skinEl = userComp.shadowRoot.querySelector('.skin-types');
            if (skinEl) {
                skinTypes = Array.from(skinEl.querySelectorAll('.skin-type'))
                                 .map(el => el.textContent.trim())
                                 .filter(Boolean)
                                 .join(', ');
            }
        }

        const key = date + '::' + content.substring(0, 40);
        if (!seen.has(key)) {
            seen.add(key);
            reviews.push({
                goods_no:    goodsNo,
                username:    username,
                skin_types:  skinTypes,
                rating:      rating,
                date:        date,
                content:     content,
                option_name: option,
            });
        }
    }
    return reviews;
})(arguments[0]);
"""

# 마지막 oy-review-review-item을 뷰포트로 스크롤 (가상 스크롤 트리거)
_SCROLL_LAST_ITEM_JS = """
(function() {
    function findAll(root, tagLower, depth) {
        if (!root || depth > 9) return [];
        let found = [];
        try {
            for (const el of Array.from(root.querySelectorAll ? root.querySelectorAll('*') : [])) {
                if (el.tagName && el.tagName.toLowerCase() === tagLower) found.push(el);
                if (el.shadowRoot) found = found.concat(findAll(el.shadowRoot, tagLower, depth + 1));
            }
        } catch(e) {}
        return found;
    }
    const items = findAll(document, 'oy-review-review-item', 0);
    if (items.length > 0) {
        items[items.length - 1].scrollIntoView({behavior: 'instant', block: 'end'});
        window.scrollBy(0, 200);
        return items.length;
    }
    window.scrollBy(0, 800);
    return 0;
})();
"""


# ==============================================================================
# API 응답 파싱 및 피부타입 코드 디코딩
# ==============================================================================

SKIN_TYPE_MAP = {
    "A01": "지성",
    "A02": "건성",
    "A03": "복합성",
    "A04": "중성",
    "A05": "약건성",
}

SKIN_TONE_MAP = {
    "B01": "웜톤",
    "B02": "쿨톤",
    "B03": "봄웜톤",
    "B04": "여름쿨톤",
    "B05": "가을웜톤",
    "B06": "겨울쿨톤",
}

SKIN_TROUBLE_MAP = {
    "C01": "민감성",
    "C02": "잡티",
    "C03": "모공",
    "C04": "각질",
    "C05": "트러블",
    "C06": "블랙헤드",
    "C07": "주름",
    "C08": "미백",
}

def decode_skin_types(profile_dto):
    """API의 profileDto에 포함된 피부타입/톤/고민 코드를 한글 텍스트로 복원."""
    if not profile_dto or not isinstance(profile_dto, dict):
        return ""
    
    parts = []
    
    # 1. 피부 타입 (skinType)
    st = profile_dto.get("skinType")
    if st:
        parts.append(SKIN_TYPE_MAP.get(st, st))
        
    # 2. 피부 톤 (skinTone)
    stone = profile_dto.get("skinTone")
    if stone:
        parts.append(SKIN_TONE_MAP.get(stone, stone))
        
    # 3. 피부 고민 (skinTrouble - 리스트)
    troubles = profile_dto.get("skinTrouble") or []
    for tr in troubles:
        if tr:
            parts.append(SKIN_TROUBLE_MAP.get(tr, tr))
            
    return ", ".join(filter(None, parts))

def parse_api_batch(body_text, goods_no):
    """인터셉트된 API 응답 JSON 한 건에서 리뷰 목록 추출."""
    reviews = []
    try:
        data = json.loads(body_text)
    except Exception:
        return reviews

    _d = data.get('data')
    _r = data.get('result')
    # 실제 API 응답: {"data": {"goodsReviewList": [...]}}
    review_list = (
        data.get('goodsReviewList') or
        data.get('reviewList') or
        (_d.get('goodsReviewList') if isinstance(_d, dict) else None) or  # ← 실제 키
        (_d.get('reviewList')      if isinstance(_d, dict) else None) or
        (_r.get('goodsReviewList') if isinstance(_r, dict) else None) or
        (_r.get('reviewList')      if isinstance(_r, dict) else None) or
        (_d if isinstance(_d, list) else None) or
        []
    )
    if not isinstance(review_list, list):
        return reviews

    for r in review_list:
        if not isinstance(r, dict):
            continue

        # 실제 필드: content (reviewContents 등도 시도)
        content = (r.get('content') or r.get('reviewContents') or
                   r.get('reviewContent') or r.get('contents') or '').strip()
        if not content or len(content) < 5:
            continue

        # 날짜: createdDateTime 최우선 적용 → 여러 필드명 시도 → 없으면 사진 경로 날짜 → reviewId 사용
        date = (r.get('createdDateTime') or r.get('reviewDt') or r.get('regDt') or r.get('date') or
                r.get('createDt') or r.get('writeDate') or r.get('createDate') or '').strip()
        if not date:
            photos = r.get('photoReviewList') or []
            if photos and isinstance(photos[0], dict):
                m = re.match(r'(\d{4})/(\d{2})/(\d{2})/', photos[0].get('imagePath', ''))
                if m:
                    date = f"{m.group(1)}.{m.group(2)}.{m.group(3)}"
        if not date:
            date = str(r.get('reviewId', ''))  # 최후 수단: ID를 날짜 대체로 사용

        # 닉네임: memberNickname(profileDto 내부) 또는 nickName 등 최우선 시도
        profile_dto = r.get('profileDto') or {}
        username = (profile_dto.get('memberNickname') or r.get('memberNickNm') or r.get('nickName') or
                    r.get('nickname') or r.get('writerNm') or r.get('memberName') or '익명').strip() or '익명'

        try:
            rating = int(r.get('reviewScore') or r.get('score') or
                         r.get('rating') or r.get('starScore') or r.get('pointScore') or 5)
        except Exception:
            rating = 5

        # 옵션: 최상위 필드 또는 goodsDto 안에 있음
        goods_dto = r.get('goodsDto') or {}
        option = (r.get('goodsOptionNm') or r.get('optionNm') or r.get('optionName') or
                  goods_dto.get('optionName') or goods_dto.get('goodsName') or '단품').strip()
        option = re.sub(r'^\[옵션\]\s*', '', option).strip() or '단품'

        review_id = str(r.get('reviewId', ''))

        # 피부타입: profileDto 내 코드 디코딩 처리
        skin_types = decode_skin_types(profile_dto)

        reviews.append({
            'review_id':   review_id,
            'goods_no':    goods_no,
            'username':    username,
            'skin_types':  skin_types,
            'rating':      max(1, min(5, rating)),
            'date':        date,
            'content':     content,
            'option_name': option,
        })

    return reviews


_cursor_body_logged = False  # /cursor 응답 본문 최초 1회만 출력

def collect_api_reviews(driver, goods_no):
    """버퍼에 쌓인 API 배치를 모두 수집하고 버퍼를 초기화."""
    global _cursor_body_logged
    raw = driver.execute_script(_COLLECT_BATCHES_JS)
    batches = json.loads(raw or '[]')
    reviews = []
    for b in batches:
        url  = b.get('url', '')
        body = b.get('body', '')
        # URL은 항상 출력 (한 줄 요약)
        short_url = url.split('?')[0].replace('https://m.oliveyoung.co.kr', '')
        print(f"    [API] {b.get('src','').upper():5} {short_url} → {b.get('status','')}")
        # /cursor 첫 응답 본문만 상세 출력 (필드명 확인용)
        if not _cursor_body_logged and '/cursor' in url and body:
            print(f"    [CURSOR body 800c] {body[:800]}")
            _cursor_body_logged = True
        reviews.extend(parse_api_batch(body, goods_no))
    return reviews, len(batches)


def extract_reviews_from_shadow_dom(driver, goods_no):
    """Shadow DOM에서 현재 렌더링된 리뷰 추출."""
    try:
        result = driver.execute_script(_EXTRACT_JS, goods_no)
        return result if result else []
    except Exception as e:
        print(f"    [!] Shadow DOM 추출 오류: {e}")
        return []


# ==============================================================================
# Selenium 드라이버 초기화
# ==============================================================================

def init_driver():
    print("\n[*] Edge 모바일 에뮬레이션 드라이버를 초기화하는 중...")
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=375,812")
    options.binary_location = EDGE_BINARY_PATH
    options.add_argument(
        "user-agent=Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) "
        "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1"
    )
    try:
        driver = webdriver.Edge(options=options)
        print("[*] 드라이버 초기화 완료 (375×812 모바일 뷰포트).")
        return driver
    except Exception as e:
        print(f"[!] 드라이버 초기화 실패: {e}")
        sys.exit(1)


# ==============================================================================
# 크롤링 보조 함수
# ==============================================================================

def _dismiss_popups(driver):
    driver.execute_script("""
        ['닫기','그만보기','웹으로','웹에서','close'].forEach(kw => {
            document.querySelectorAll('button,a,span').forEach(el => {
                if (el.textContent.trim().toLowerCase().includes(kw) && el.offsetWidth > 0)
                    try { el.click(); } catch(e) {}
            });
        });
        ['[class*="popup"]','[class*="modal"]','[class*="dimmed"]',
         '[class*="overlay"]','[class*="layer"]','[class*="banner"]'].forEach(sel =>
            document.querySelectorAll(sel).forEach(el => {
                const c = (el.className||'').toLowerCase(), id = (el.id||'').toLowerCase();
                if (!c.includes('review') && !c.includes('gdas') && !id.includes('review'))
                    el.style.display = 'none';
            })
        );
        document.documentElement.style.overflow = 'auto';
        document.body.style.overflow  = 'auto';
        document.body.style.position  = 'static';
    """)


def _activate_review_tab(driver):
    """
    '리뷰' 탭 클릭 후 oy-review-review-in-product 마운트를 폴링으로 확인.
    마운트가 확인될 때까지 최대 20초 대기.
    """
    _SKIP_CLS = ("ReviewArea_btn-review", "ReviewArea_review-thumbs",
                 "review-count", "review-score")

    def _try_click(el):
        try:
            txt = el.text.strip()
            cls = el.get_attribute("class") or ""
            tag = el.tag_name.lower()
            if ('리뷰' in txt and 1 <= len(txt) <= 40
                    and tag in ('li', 'button', 'a', 'span', 'div', 'em')
                    and not any(s in cls for s in _SKIP_CLS)):
                driver.execute_script("arguments[0].click();", el)
                print(f"[*] 탭 클릭 시도: <{tag}> cls='{cls[:40]}' txt='{txt[:30]}'")
                return True
        except Exception:
            pass
        return False

    # 1순위: GoodsDetailTabs 내부 탭 항목 (첫 성공 즉시 탈출)
    clicked = False
    for selector in ["[class*='GoodsDetailTabs'] li",
                     "[class*='GoodsDetailTabs'] button",
                     "li[role='tab']", "button[role='tab']",
                     ".tab_menu li", "[class*='tab-menu'] li",
                     "[class*='detail-tab'] li"]:
        for el in driver.find_elements(By.CSS_SELECTOR, selector):
            if _try_click(el):
                clicked = True
                break
        if clicked:
            break

    # 2순위: XPath (CSS 선택자로 못 찾은 경우에만)
    if not clicked:
        for el in driver.find_elements(By.XPATH, "//*[contains(text(), '리뷰')]"):
            if _try_click(el):
                break

    # ── 컴포넌트 마운트 폴링 (최대 20초) ─────────────────────────────
    for i in range(20):
        time.sleep(1.0)
        mounted = driver.execute_script(
            "return !!document.querySelector('oy-review-review-in-product');"
        )
        if mounted:
            print(f"[*] oy-review-review-in-product 마운트 확인 ({i+1}초 경과)")
            return True

    # 마운트 실패 — 진단 정보 출력
    oy_tags = driver.execute_script("""
        return [...new Set(
            Array.from(document.querySelectorAll('*'))
                .filter(e => e.tagName.toLowerCase().startsWith('oy-'))
                .map(e => e.tagName.toLowerCase())
        )].slice(0, 20);
    """)
    tabs_found = driver.execute_script("""
        return Array.from(document.querySelectorAll('*'))
            .filter(e => {
                const t = (e.textContent || '').trim();
                return t.includes('리뷰') && t.length < 30 && e.children.length === 0;
            })
            .map(e => ({tag: e.tagName, cls: (e.className||'').substring(0,50),
                        txt: e.textContent.trim().substring(0,20),
                        y: e.getBoundingClientRect().top}))
            .slice(0, 10);
    """)
    print(f"[!] 마운트 실패 — oy-* 컴포넌트: {oy_tags}")
    print(f"[!] '리뷰' 텍스트 요소: {tabs_found}")
    return False


def _scroll_to_review_component(driver):
    """oy-review-review-in-product 컴포넌트가 뷰포트에 들어오도록 스크롤."""
    driver.execute_script("""
        const comp = document.querySelector('oy-review-review-in-product');
        if (comp) comp.scrollIntoView({behavior: 'instant', block: 'start'});
        else window.scrollTo(0, document.body.scrollHeight * 0.4);
    """)


# ==============================================================================
# 옵션 필터링 기반 크롤링 함수
# ==============================================================================

def crawl_with_options_filtering(driver, goods_no, page_info):
    """
    [v5 추가 기능]
    상품 상세 페이지에서 '상품 옵션' 필터링 버튼을 통해 모달을 호출하고,
    각 옵션(예: 각 패드 기획 상품)별로 리뷰 페이지를 필터링 조회하여 수집하는 고유 기능.
    """
    url = f"https://m.oliveyoung.co.kr/m/goods/getGoodsDetail.do?goodsNo={goods_no}"
    print(f"\n{'='*60}")
    print(f"[*] [옵션 필터링 적용] 크롤링 시작: [{goods_no}] {page_info['name']}")
    print(f"[*] URL: {url}")

    driver.get(url)
    time.sleep(5.5)

    # 인터셉터 주입 (리뷰 탭 클릭 전 — 이후 API 요청 전부 캡처)
    driver.execute_script(_INTERCEPTOR_JS)
    time.sleep(0.3)
    print("[*] API 인터셉터 주입 완료.")

    _dismiss_popups(driver)
    time.sleep(1.0)

    mounted = _activate_review_tab(driver)
    if not mounted:
        print("[!] 리뷰 컴포넌트 마운트 실패.")
        return []

    _scroll_to_review_component(driver)
    time.sleep(2.0)

    # 1. Click "상품 옵션" button inside the shadowRoot of oy-review-filter-chips to open the option list
    open_success = driver.execute_script("""
        function findDeep(root, tag) {
            if (!root) return null;
            const el = root.querySelector ? root.querySelector(tag) : null;
            if (el) return el;
            for (const e of (root.querySelectorAll ? Array.from(root.querySelectorAll('*')) : [])) {
                if (e.shadowRoot) { const r = findDeep(e.shadowRoot, tag); if (r) return r; }
            }
            return null;
        }
        const comp = document.querySelector('oy-review-review-in-product');
        if (!comp || !comp.shadowRoot) return false;
        const chipsComp = findDeep(comp.shadowRoot, 'oy-review-filter-chips');
        if (!chipsComp || !chipsComp.shadowRoot) return false;
        const list = Array.from(chipsComp.shadowRoot.querySelectorAll('li'));
        for (const li of list) {
            if (li.textContent.trim().includes('상품 옵션')) {
                const btn = li.querySelector('oy-review-common-button');
                if (btn && btn.shadowRoot) {
                    const actualBtn = btn.shadowRoot.querySelector('button');
                    if (actualBtn) { actualBtn.click(); return true; }
                }
                li.click();
                return true;
            }
        }
        return false;
    """)
    print(f"[*] 상품 옵션 모달 오픈 시도: {open_success}")
    time.sleep(2.5)

    # 2. Retrieve all options from oy-review-goods-option-sheet shadow DOM
    options_list = driver.execute_script("""
        function findDeep(root, tag) {
            if (!root) return null;
            const el = root.querySelector ? root.querySelector(tag) : null;
            if (el) return el;
            for (const e of (root.querySelectorAll ? Array.from(root.querySelectorAll('*')) : [])) {
                if (e.shadowRoot) { const r = findDeep(e.shadowRoot, tag); if (r) return r; }
            }
            return null;
        }
        const sheet = findDeep(document, 'oy-review-goods-option-sheet');
        if (!sheet || !sheet.shadowRoot) return [];
        const options = Array.from(sheet.shadowRoot.querySelectorAll('li.option'));
        return options.map((opt, idx) => {
            const name = opt.querySelector('.option-name').textContent.trim();
            const count = opt.querySelector('.review-count').textContent.trim();
            return { index: idx, name: name, count: count };
        });
    """)

    print(f"[*] 총 {len(options_list)}개의 옵션을 탐지했습니다:")
    for opt in options_list:
        print(f"    - [{opt['index']}] {opt['name']} ({opt['count']})")

    all_option_reviews = {}

    for opt_idx, opt in enumerate(options_list):
        opt_name = opt['name']
        opt_count_str = opt['count']
        
        # Parse count to integer
        parsed_count = 0
        m = re.search(r'(\d+)', opt_count_str.replace(',', ''))
        if m:
            parsed_count = int(m.group(1))
        
        # Determine dynamic target for this option
        opt_target = min(250, parsed_count)
        if opt_target == 0:
            print(f"\n[~] [{opt_idx+1}/{len(options_list)}] {opt_name} -> 리뷰 0건. 건너뜁니다.")
            continue
            
        print(f"\n{'-'*50}")
        print(f"[+] [{opt_idx+1}/{len(options_list)}] 옵션 필터 선택: {opt_name}")
        print(f"[+] 해당 옵션 총 리뷰 수: {parsed_count}건, 수집 목표: {opt_target}개")

        # Open options modal if not open
        is_open = driver.execute_script("""
            function findDeep(root, tag) {
                if (!root) return null;
                const el = root.querySelector ? root.querySelector(tag) : null;
                if (el) return el;
                for (const e of (root.querySelectorAll ? Array.from(root.querySelectorAll('*')) : [])) {
                    if (e.shadowRoot) { const r = findDeep(e.shadowRoot, tag); if (r) return r; }
                }
                return null;
            }
            const comp = findDeep(document, 'oy-review-bottom-sheet');
            if (comp && comp.shadowRoot) {
                const container = comp.shadowRoot.querySelector('.bottom-sheet-container');
                return container ? container.classList.contains('is-open') : false;
            }
            return false;
        """)
        
        if not is_open:
            print("[*] 옵션 필터 모달이 닫혀 있어 다시 엽니다...")
            driver.execute_script("""
                function findDeep(root, tag) {
                    if (!root) return null;
                    const el = root.querySelector ? root.querySelector(tag) : null;
                    if (el) return el;
                    for (const e of (root.querySelectorAll ? Array.from(root.querySelectorAll('*')) : [])) {
                        if (e.shadowRoot) { const r = findDeep(e.shadowRoot, tag); if (r) return r; }
                    }
                    return null;
                }
                const comp = document.querySelector('oy-review-review-in-product');
                const chipsComp = findDeep(comp.shadowRoot, 'oy-review-filter-chips');
                const list = Array.from(chipsComp.shadowRoot.querySelectorAll('li'));
                for (const li of list) {
                    if (li.textContent.trim().includes('상품 옵션')) {
                        const btn = li.querySelector('oy-review-common-button');
                        if (btn && btn.shadowRoot) {
                            btn.shadowRoot.querySelector('button').click();
                        } else {
                            li.click();
                        }
                        break;
                    }
                }
            """)
            time.sleep(2.0)

        # Select option in UI
        driver.execute_script("""
            const targetIndex = arguments[0];
            function findDeep(root, tag) {
                if (!root) return null;
                const el = root.querySelector ? root.querySelector(tag) : null;
                if (el) return el;
                for (const e of (root.querySelectorAll ? Array.from(root.querySelectorAll('*')) : [])) {
                    if (e.shadowRoot) { const r = findDeep(e.shadowRoot, tag); if (r) return r; }
                }
                return null;
            }
            const sheet = findDeep(document, 'oy-review-goods-option-sheet');
            
            // 1. Click reset button first to clear existing selections
            const resetBtn = sheet.shadowRoot.querySelector('.reset-button');
            if (resetBtn) resetBtn.click();
            
            // 2. Select the option at targetIndex
            const options = Array.from(sheet.shadowRoot.querySelectorAll('li.option'));
            if (options[targetIndex]) {
                options[targetIndex].click();
            }
            
            // 3. Click review-button
            const viewBtn = sheet.shadowRoot.querySelector('.review-button');
            if (viewBtn) viewBtn.click();
        """, opt_idx)
        
        time.sleep(2.5)  # Wait for reviews to refresh

        option_reviews = {}
        
        # ── 1차 응답 수집: 필터 적용 직후 로드된 리뷰 수집 (특히 리뷰 수가 적은 희소 상품용) ──
        api_reviews, batch_count = collect_api_reviews(driver, goods_no)
        dom_reviews = extract_reviews_from_shadow_dom(driver, goods_no)
        for r in api_reviews + dom_reviews:
            key = r.get('review_id') or (r['date'] + '::' + r['content'][:40])
            option_reviews[key] = r

        # Initialize scroll variables for this option
        max_scrolls = page_info["max_scroll_steps"]
        delay = page_info["delay"]
        MAX_NO_GROWTH = 15  # Slightly shorter for per-option crawl to save time
        
        last_count = len(option_reviews)
        no_growth = 0
        
        print(f"[*] [{opt_name}] 옵션 리뷰 스크롤 및 크롤링 시작...")

        for step in range(1, max_scrolls + 1):
            # ── 1) API 인터셉터에서 새 배치 수집 ──────────────────────
            api_reviews, batch_count = collect_api_reviews(driver, goods_no)
            
            # ── 2) Shadow DOM에서 현재 렌더링된 아이템 추출 ───────────
            dom_reviews = extract_reviews_from_shadow_dom(driver, goods_no)
            
            # ── 3) 중복 제거 병합 ─────────────────────────────────────
            for r in api_reviews + dom_reviews:
                key = r.get('review_id') or (r['date'] + '::' + r['content'][:40])
                option_reviews[key] = r
                
            current = len(option_reviews)
            growth = current - last_count
            print(f"  [{step:>4}/{max_scrolls}] 옵션고유: {current:>5}/{opt_target} (+{growth:>3}) API:{batch_count} DOM:{len(dom_reviews)}")
            
            if current >= opt_target:
                print(f"  [+] 옵션 목표 {opt_target}개 달성! 조기 종료.")
                break
                
            if current == last_count:
                no_growth += 1
                if no_growth >= MAX_NO_GROWTH:
                    print(f"  [!] {MAX_NO_GROWTH}회 연속 미증가. 옵션 종료.")
                    break
                if no_growth % 5 == 0:
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(3.0)
            else:
                no_growth = 0
                
            last_count = current
            
            # scrollIntoView to trigger virtual scroll IntersectionObserver
            rendered = driver.execute_script(_SCROLL_LAST_ITEM_JS)
            if rendered == 0:
                driver.execute_script("window.scrollBy(0, 800);")
                
            time.sleep(delay + random.uniform(0.1, 0.3))
            
        print(f"[+] [{opt_name}] 최종 {len(option_reviews)}개 고유 리뷰 수집 완료.")
        all_option_reviews.update(option_reviews)

    print(f"\n{'-' * 60}")
    print(f"[*] [{goods_no}] 통합 옵션 필터링 완료: 총 {len(all_option_reviews)}개 고유 리뷰 수집됨.")
    print(f"{'=' * 60}\n")
    return list(all_option_reviews.values())


# ==============================================================================
# 메인 크롤링 함수
# ==============================================================================

def crawl_raw_reviews_from_page(driver, goods_no, page_info):
    """
    [v4 변경 사항]
    1. XHR + fetch 인터셉터 → /review/api/v2/reviews/cursor 응답 JSON 직접 수집
       (가상 스크롤 DOM에 7개만 렌더링되는 문제 우회)
    2. Shadow DOM 추출 병행 (인터셉터 파싱 실패 시 보완)
    3. scrollBy 대신 마지막 oy-review-review-item.scrollIntoView 사용
       → 가상 스크롤 IntersectionObserver 정확히 트리거
    """
    if goods_no in ["A000000166709", "A000000206889"]:
        return crawl_with_options_filtering(driver, goods_no, page_info)

    url = f"https://m.oliveyoung.co.kr/m/goods/getGoodsDetail.do?goodsNo={goods_no}"
    print(f"\n{'='*60}")
    print(f"[*] 크롤링 시작: [{goods_no}] {page_info['name']}")
    print(f"[*] URL: {url}")
    print(f"[*] 목표: {page_info['target_reviews']}개, 최대 스크롤: {page_info['max_scroll_steps']}회")

    driver.get(url)
    time.sleep(5.5)

    # 인터셉터 주입 (리뷰 탭 클릭 전 — 이후 API 요청 전부 캡처)
    driver.execute_script(_INTERCEPTOR_JS)
    time.sleep(0.3)
    print("[*] API 인터셉터 주입 완료.")

    _dismiss_popups(driver)
    time.sleep(1.0)

    _activate_review_tab(driver)
    # _activate_review_tab이 내부에서 마운트 폴링으로 대기하므로 별도 sleep 최소화
    time.sleep(2.0)  # 1차 API 응답 수신 여유 시간

    _scroll_to_review_component(driver)
    time.sleep(2.0)

    # 초기 스크롤: 리뷰 아이템이 뷰포트에 진입하도록 먼저 내려감
    print("[*] 초기 스크롤 (5×700px) — virtual scroll 렌더링 대기...")
    for _ in range(5):
        driver.execute_script("window.scrollBy(0, 700);")
        time.sleep(1.2)
    time.sleep(2.0)

    # ── 최초 Shadow DOM 상태 진단 ────────────────────────────────
    diag = driver.execute_script("""
        function findDeep(root, tag) {
            if (!root) return null;
            const el = root.querySelector ? root.querySelector(tag) : null;
            if (el) return el;
            for (const e of (root.querySelectorAll ? Array.from(root.querySelectorAll('*')) : [])) {
                if (e.shadowRoot) { const r = findDeep(e.shadowRoot, tag); if (r) return r; }
            }
            return null;
        }
        const comp = document.querySelector('oy-review-review-in-product');
        if (!comp) return {step: 'no-comp'};
        if (!comp.shadowRoot) return {step: 'no-comp-shadow'};
        const list = findDeep(comp.shadowRoot, 'oy-review-review-list');
        if (!list) return {step: 'no-list', compHTML: comp.shadowRoot.innerHTML.substring(0, 300)};
        if (!list.shadowRoot) return {step: 'no-list-shadow'};
        const items = Array.from(list.shadowRoot.querySelectorAll('oy-review-review-item'));
        const allTags = [...new Set(Array.from(list.shadowRoot.querySelectorAll('*'))
            .map(e => e.tagName.toLowerCase()))].slice(0, 20);
        return {step: 'ok', itemCount: items.length, allTags,
                listHTML: list.shadowRoot.innerHTML.substring(0, 400)};
    """)
    print(f"[*] Shadow DOM 초기 진단: {json.dumps(diag, ensure_ascii=False)}")

    target        = page_info["target_reviews"]
    max_scrolls   = page_info["max_scroll_steps"]
    delay         = page_info["delay"]
    MAX_NO_GROWTH = 30

    raw_reviews   = {}
    last_count    = 0
    no_growth     = 0

    for step in range(1, max_scrolls + 1):
        # ── 1) API 인터셉터에서 새 배치 수집 ──────────────────────
        api_reviews, batch_count = collect_api_reviews(driver, goods_no)

        # ── 2) Shadow DOM에서 현재 렌더링된 아이템 추출 ───────────
        dom_reviews = extract_reviews_from_shadow_dom(driver, goods_no)

        # ── 3) 중복 제거 병합 ─────────────────────────────────────
        for r in api_reviews + dom_reviews:
            # review_id가 있으면 그걸로, 없으면 날짜+내용 앞글자로 키 생성
            key = r.get('review_id') or (r['date'] + '::' + r['content'][:40])
            raw_reviews[key] = r

        current = len(raw_reviews)
        growth  = current - last_count
        print(f"  [{step:>4}/{max_scrolls}] 고유: {current:>5}/{target}"
              f"  (+{growth:>3})  API배치:{batch_count}({len(api_reviews)}건)  DOM:{len(dom_reviews)}개")

        if current >= target:
            print(f"  [+] 목표 {target}개 달성! 조기 종료.")
            break

        if current == last_count:
            no_growth += 1
            if no_growth >= MAX_NO_GROWTH:
                print(f"  [!] {MAX_NO_GROWTH}회 연속 미증가. 종료.")
                break
            if no_growth % 10 == 0:
                # 리뷰 컴포넌트 상단으로 돌아간 뒤 다시 내려가기
                _scroll_to_review_component(driver)
                time.sleep(2.0)
                print(f"  [~] no_growth={no_growth}: 컴포넌트 상단으로 재위치")
            elif no_growth % 5 == 0:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                print(f"  [~] no_growth={no_growth}: scrollTo(bottom) 강제 실행")
                time.sleep(4.0)
                continue
        else:
            no_growth = 0

        last_count = current

        # ── 마지막 렌더링 아이템 scrollIntoView → 가상 스크롤 트리거 ──
        rendered = driver.execute_script(_SCROLL_LAST_ITEM_JS)
        if rendered == 0:
            # 아이템 없음 — 단순 scrollBy 폴백
            driver.execute_script("window.scrollBy(0, 800);")

        time.sleep(delay + random.uniform(0.2, 0.5))

    print(f"[*] [{goods_no}] 최종 {len(raw_reviews)}개 고유 리뷰 수집 완료.")
    return list(raw_reviews.values())


# ==============================================================================
# 리뷰 분류
# ==============================================================================

def map_reviews_to_target_pads(all_raw_reviews):
    print("\n" + "="*60)
    print("[*] 11종 패드로 리뷰 분류 시작...")

    categorized = {name: [] for name in TARGET_PADS}
    unmapped    = 0

    for r in all_raw_reviews:
        mapped = False
        opt    = (r['option_name'] or "").lower()
        gno    = r['goods_no']

        for pad, info in TARGET_PADS.items():
            for kw in info["keywords"]:
                if kw in opt:
                    categorized[pad].append(r)
                    mapped = True
                    break
            if mapped:
                break

        if not mapped:
            for pad, info in TARGET_PADS.items():
                if gno == info["default_goods"]:
                    categorized[pad].append(r)
                    mapped = True
                    break

        if not mapped:
            unmapped += 1

    print("[*] 분류 결과:")
    for pad, revs in categorized.items():
        print(f"    - {pad}: {len(revs)}개")
    print(f"    - 미분류: {unmapped}개")
    return categorized


# ==============================================================================
# 균등 평점 샘플링
# ==============================================================================

def balance_reviews(pad_name, reviews, target_total=185):
    print(f"\n[*] [{pad_name}] 평점 균등 샘플링...")

    by_rating   = {r: [] for r in range(1, 6)}
    for rev in reviews:
        s = rev["rating"]
        if s in by_rating:
            by_rating[s].append(rev)

    available   = {r: len(by_rating[r]) for r in range(1, 6)}
    total_avail = sum(available.values())
    if total_avail <= target_total:
        print(f"    [!] 가용({total_avail}) ≤ 목표({target_total}). 전체 반환.")
        return reviews

    allocated = {r: 0 for r in range(1, 6)}
    remaining = target_total
    active    = list(range(1, 6))

    while remaining > 0 and active:
        share      = max(remaining // len(active), 1)
        next_active = []
        for r in active:
            avail = available[r] - allocated[r]
            give  = min(avail, share)
            allocated[r] += give
            remaining    -= give
            if avail > share:
                next_active.append(r)
        if len(next_active) == len(active) and remaining > 0:
            for r in sorted(next_active, key=lambda x: available[x]-allocated[x], reverse=True):
                if remaining > 0 and (available[r]-allocated[r]) > 0:
                    allocated[r] += 1
                    remaining    -= 1
            break
        active = next_active

    selected = []
    for r in range(1, 6):
        pool = by_rating[r][:]
        random.shuffle(pool)
        selected.extend(pool[:allocated[r]])
        print(f"    - {r}점: {available[r]}개 중 {allocated[r]}개 선택")

    print(f"    → 선정 완료: {len(selected)}개")
    return selected


# ==============================================================================
# 엑셀 저장
# ==============================================================================

def save_to_premium_excel(final_reviews_dict):
    print(f"\n[*] 엑셀 저장 준비... ({OUTPUT_FILENAME})")

    # 기존에 파일이 이미 저장되어 있다면, 저장된 데이터를 삭제
    if os.path.exists(OUTPUT_FILENAME):
        try:
            os.remove(OUTPUT_FILENAME)
            print(f"[*] 기존의 '{OUTPUT_FILENAME}' 파일 및 데이터를 성공적으로 삭제했습니다. 새로운 데이터로 갱신합니다.")
        except Exception as e:
            print(f"[!] 기존 파일 삭제 중 오류 발생 (파일이 열려있을 수 있습니다): {e}")

    rows = []
    for pad_name, revs in final_reviews_dict.items():
        for r in revs:
            rows.append({
                "타겟상품명":        pad_name,
                "올리브영 상품코드": r["goods_no"],
                "구매 옵션명":       r["option_name"] or "단품",
                "작성자":            r["username"],
                "피부타입":          r["skin_types"],
                "별점":              f"{r['rating']}점",
                "작성일":            r["date"],
                "리뷰 내용":         r["content"],
            })

    df = pd.DataFrame(rows)
    if df.empty:
        print("[!] 저장할 데이터 없음.")
        return

    cols = ["타겟상품명", "올리브영 상품코드", "구매 옵션명", "작성자", "피부타입", "별점", "작성일", "리뷰 내용"]
    df   = df[cols]

    try:
        with pd.ExcelWriter(OUTPUT_FILENAME, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='스킨푸드 패드 리뷰')
            ws  = writer.sheets['스킨푸드 패드 리뷰']

            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            hdr_fill  = PatternFill(start_color="33691E", end_color="33691E", fill_type="solid")
            hdr_font  = Font(name="Malgun Gothic", size=11, bold=True, color="FFFFFF")
            body_font = Font(name="Malgun Gothic", size=10)
            thin      = Border(**{s: Side(style='thin', color='EAEAEA')
                                   for s in ('left', 'right', 'top', 'bottom')})

            for col in range(1, len(cols) + 1):
                c = ws.cell(row=1, column=col)
                c.fill = hdr_fill; c.font = hdr_font; c.border = thin
                c.alignment = Alignment(horizontal="center", vertical="center")

            for row in range(2, len(df) + 2):
                for col in range(1, len(cols) + 1):
                    c = ws.cell(row=row, column=col)
                    c.font = body_font; c.border = thin
                    if col in (1, 2, 4, 6, 7):
                        c.alignment = Alignment(horizontal="center", vertical="center")
                    elif col == 8:
                        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    else:
                        c.alignment = Alignment(horizontal="left", vertical="center")

            ws.row_dimensions[1].height = 32
            for r in range(2, len(df) + 2):
                ws.row_dimensions[r].height = 25

            for col in ws.columns:
                letter  = col[0].column_letter
                max_len = max(
                    ((len((v := str(c.value or '')).encode('utf-8')) + len(v)) // 2)
                    for c in col
                )
                ws.column_dimensions[letter].width = min(max(max_len + 4, 12), 70)

        print(f"\n[+] 저장 완료: {os.path.abspath(OUTPUT_FILENAME)}")
        print(f"[+] 총 {len(df)}개 행 저장됨.")
    except Exception as e:
        print(f"[!] 엑셀 저장 오류: {e}")


# ==============================================================================
# 진입점
# ==============================================================================

def main():
    print("="*60)
    print("  올리브영 스킨푸드 11종 패드 균등 평점 리뷰 수집기 v4")
    print("  방식: API 인터셉터 + Shadow DOM 추출 + 가상 스크롤 트리거")
    print("="*60)

    driver = init_driver()
    all_reviews = []

    try:
        for idx, (goods_no, page_info) in enumerate(PRODUCT_PAGES.items()):
            print(f"\n[페이지 {idx+1}/{len(PRODUCT_PAGES)}]")
            page_reviews = crawl_raw_reviews_from_page(driver, goods_no, page_info)
            all_reviews.extend(page_reviews)

            if idx < len(PRODUCT_PAGES) - 1:
                print("[*] 다음 페이지 이동 전 4초 쿨다운...")
                time.sleep(4)
    finally:
        print("\n[*] Selenium 드라이버 종료.")
        driver.quit()

    categorized = map_reviews_to_target_pads(all_reviews)

    final = {}
    for pad, pool in categorized.items():
        final[pad] = balance_reviews(pad, pool, target_total=185)

    save_to_premium_excel(final)

    print("\n" + "="*60)
    print("  모든 작업 완료!")
    print("="*60)


if __name__ == "__main__":
    main()
