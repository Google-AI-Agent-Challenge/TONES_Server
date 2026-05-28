import time
import re
import sys
import os
import random
import pandas as pd
from bs4 import BeautifulSoup
from curl_cffi import requests
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By

# 콘솔 출력 인코딩을 UTF-8로 강제 설정하여 한글 및 이모지 깨짐 방지
sys.stdout.reconfigure(encoding='utf-8')

# ==============================================================================
# CONFIGURATION BLOCK (설정 영역)
# ==============================================================================
OUTPUT_FILENAME = "스킨푸드_패드_고객리뷰.xlsx"
EDGE_BINARY_PATH = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
# ==============================================================================

# 스킨푸드 11종 패드 상품 정의 및 옵션 매핑 키워드 규칙
TARGET_PADS = {
    "아스파라거스 패드": {
        "keywords": ["아스파라거스", "asparagus"],
        "default_goods": "A000000166709"
    },
    "복숭아 패드": {
        "keywords": ["복숭아", "피치", "peach"],
        "default_goods": "A000000231714"
    },
    "블루 캐모마일 패드": {
        "keywords": ["캐모마일", "chamomile", "카모마일", "블루캐모마일"],
        "default_goods": "A000000166709"
    },
    "라이스 패드": {
        "keywords": ["라이스", "rice", "쌀"],
        "default_goods": "A000000166709"
    },
    "레몬그라스 패드": {
        "keywords": ["레몬그라스", "lemongrass"],
        "default_goods": "A000000166709"
    },
    "샤인머스캣 패드": {
        "keywords": ["샤인머스캣", "shine", "머스캣"],
        "default_goods": "A000000166709"
    },
    "핑크자몽 패드": {
        "keywords": ["자몽", "grapefruit", "핑크자몽"],
        "default_goods": "A000000166709"
    },
    "미나리 패드": {
        "keywords": ["미나리", "파슬리", "parsley", "판토텐산"],
        "default_goods": "A000000185135"
    },
    "당근 패드": {
        "keywords": ["당근", "캐롯", "carrot"],
        "default_goods": "A000000248098"
    },
    "감자 패드": {
        "keywords": ["감자", "포테이토", "potato"],
        "default_goods": "A000000200396"
    },
    "도토리 패드": {
        "keywords": ["도토리", "에이콘", "acorn"],
        "default_goods": "A000000157075"
    }
}

# 고유 상품 페이지 정의 및 수집할 스크롤 스텝 분기
# 통합 기획전 페이지일수록 더 많은 리뷰 풀을 모으기 위해 스크롤을 길게 줍니다.
PRODUCT_PAGES = {
    "A000000166709": {"name": "11종 통합 기획전 페이지 (아스파라거스/라이스/도토리 등)", "scroll_steps": 500, "delay": 1.3},
    "A000000231714": {"name": "복숭아 패드 전용/기획 페이지", "scroll_steps": 150, "delay": 1.0},
    "A000000185135": {"name": "미나리 패드 전용 페이지", "scroll_steps": 150, "delay": 1.0},
    "A000000248098": {"name": "당근 패드 기획전 페이지", "scroll_steps": 150, "delay": 1.0},
    "A000000200396": {"name": "감자 패드 전용 페이지", "scroll_steps": 150, "delay": 1.0},
    "A000000157075": {"name": "도토리 패드 전용 페이지", "scroll_steps": 150, "delay": 1.0}
}

# 중첩된 웹 컴포넌트(Shadow Root) 내부에 숨겨진 리뷰 데이터를 깊이 탐색하여 추출하는 핵심 JavaScript 코드
JS_DEEP_EXTRACT = """
let inProd = document.querySelector('oy-review-review-in-product');
if (!inProd) return [];
let root1 = inProd.shadowRoot;
if (!root1) return [];

let listProvider = root1.querySelector('oy-review-review-list-provider');
if (!listProvider) return [];
let reviewList = listProvider.querySelector('oy-review-review-list');
if (!reviewList) return [];
let root2 = reviewList.shadowRoot;
if (!root2) return [];

let items = root2.querySelectorAll('oy-review-review-item');
let reviews = [];

items.forEach(item => {
    let itemRoot = item.shadowRoot;
    if (!itemRoot) return;
    
    // 1. 작성자 닉네임 및 피부정보 추출
    let username = "익명";
    let skinTypes = [];
    let userEl = itemRoot.querySelector('oy-review-review-user');
    if (userEl && userEl.shadowRoot) {
        let nameEl = userEl.shadowRoot.querySelector('.name');
        if (nameEl) username = nameEl.textContent.trim();
        
        let skinEls = userEl.shadowRoot.querySelectorAll('.skin-type');
        skinEls.forEach(s => skinTypes.push(s.textContent.trim()));
    }
    
    // 2. 별점 추출 (별 아이콘 컴포넌트 갯수 세기)
    let rating = 0;
    let ratingEl = itemRoot.querySelector('.rating');
    if (ratingEl) {
        rating = ratingEl.querySelectorAll('oy-review-star-icon').length;
    }
    
    // 3. 작성일 추출
    let date = "";
    let dateEl = itemRoot.querySelector('.common-info .date');
    if (dateEl) date = dateEl.textContent.trim();
    
    // 4. 리뷰 상세 본문 추출
    let content = "";
    let contentEl = itemRoot.querySelector('oy-review-review-content');
    if (contentEl && contentEl.shadowRoot) {
        let pEl = contentEl.shadowRoot.querySelector('.content p');
        if (pEl) content = pEl.textContent.trim();
    }
    
    // 5. 구매 옵션명 추출
    let optionName = "";
    let optionEl = itemRoot.querySelector('.goods-option');
    if (optionEl) {
        optionName = optionEl.textContent.trim();
    } else {
        let optEl = itemRoot.querySelector('.option') || itemRoot.querySelector('[class*="option"]');
        if (optEl) optionName = optEl.textContent.trim();
    }
    
    if (username || content) {
        reviews.push({
            username: username,
            skin_types: skinTypes.join(', '),
            rating: rating,
            date: date,
            content: content,
            option_name: optionName
        });
    }
});

return reviews;
"""

JS_PAGINATE_NEXT = """
const targetPage = arguments[0];

// ---------------------------------------------------------------
// STEP 1: 리뷰 컴포넌트 Shadow DOM 내부를 우선적으로 탐색
// 리뷰 아이템 관련 태그는 탐색에서 완전히 제외
// (리뷰 본문 안에 '다음'이라는 단어가 있으면 오클릭되는 문제 방지)
// ---------------------------------------------------------------
const SKIP_TAGS = new Set([
    'OY-REVIEW-REVIEW-ITEM', 'OY-REVIEW-REVIEW-CONTENT',
    'OY-REVIEW-REVIEW-USER', 'OY-REVIEW-STAR-ICON'
]);

function searchInRoot(root, depth) {
    if (!root || depth > 6) return null;
    let els = Array.from(root.querySelectorAll('*'));
    
    // 1차: 숫자 페이지 버튼 검색 (리뷰 아이템 태그 제외)
    for (let el of els) {
        if (SKIP_TAGS.has(el.tagName)) continue;
        let text = el.textContent.trim();
        let cls = String(el.className || '').toLowerCase();
        if (text === targetPage.toString() &&
            !cls.includes('swiper') && !cls.includes('slide') && !cls.includes('banner')) {
            return { el, text, tag: el.tagName, cls, source: 'numbered-in-review' };
        }
    }
    
    // 2차: 다음/next 버튼 검색 (리뷰 아이템 태그와 해당 shadow root 순회 제외)
    for (let el of els) {
        if (SKIP_TAGS.has(el.tagName)) continue;
        let text = el.textContent.trim();
        let cls = String(el.className || '').toLowerCase();
        let ariaLabel = (el.getAttribute ? (el.getAttribute('aria-label') || '') : '').toLowerCase();
        let isNext = (
            text === '>' || text === '다음' || text === '>>' || text === '»' ||
            cls.includes('btn-next') || cls.includes('next') ||
            ariaLabel.includes('다음') || ariaLabel.includes('next')
        );
        if (isNext && !cls.includes('swiper') && !cls.includes('slide') && !cls.includes('banner')) {
            return { el, text, tag: el.tagName, cls, source: 'next-in-review' };
        }
        // SKIP_TAGS가 아닌 코알드만 shadow root 순회
        if (el.shadowRoot && !SKIP_TAGS.has(el.tagName)) {
            let found = searchInRoot(el.shadowRoot, depth + 1);
            if (found) return found;
        }
    }
    return null;
}

// 리뷰 컴포넌트 내부 우선 탐색
let inProd = document.querySelector('oy-review-review-in-product');
if (inProd && inProd.shadowRoot) {
    let result = searchInRoot(inProd.shadowRoot, 0);
    if (result) {
        result.el.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true, view: window }));
        return { clicked: true, text: result.text, tag: result.tag, cls: result.cls, source: result.source };
    }
}

// ---------------------------------------------------------------
// STEP 2: 전체 Shadow DOM 폴백 스캔 (SVGAnimatedString 안전 처리)
// ---------------------------------------------------------------
function collectAllElements(root, results) {
    if (!root) return;
    let els = Array.from(root.querySelectorAll('*'));
    for (let el of els) {
        results.push(el);
        if (el.shadowRoot && !SKIP_TAGS.has(el.tagName)) collectAllElements(el.shadowRoot, results);
    }
}

function getAncestorPath(el) {
    let path = '';
    let cur = el;
    let depth = 0;
    while (cur && depth < 15) {
        path += ' ' + (cur.tagName || '') + '.' + String(cur.className || '') + '|' + (cur.id || '');
        if (cur.parentNode) cur = cur.parentNode;
        else if (cur.host) cur = cur.host;
        else break;
        depth++;
    }
    return path.toLowerCase();
}

let allElements = [];
collectAllElements(document, allElements);

let nextCandidates = [];
for (let el of allElements) {
    if (SKIP_TAGS.has(el.tagName)) continue;
    let text = el.textContent.trim();
    let cls = String(el.className || '').toLowerCase();
    let id = (el.id || '').toLowerCase();
    let ariaLabel = (el.getAttribute ? (el.getAttribute('aria-label') || '') : '').toLowerCase();
    let isNextBtn = (
        text === '>' || text === '다음' || text === 'next' || text === '>>' || text === '»' ||
        cls.includes('next') || id.includes('next') ||
        ariaLabel.includes('next') || ariaLabel.includes('다음')
    );
    if (!isNextBtn) continue;
    let path = getAncestorPath(el);
    if (path.includes('swiper') || path.includes('carousel') ||
        path.includes('visual') || path.includes('slider') ||
        path.includes('slide') || path.includes('banner')) continue;
    let score = 0;
    if (path.includes('review') && path.includes('pagination')) score += 40;
    if (path.includes('oy-review')) score += 30;
    if (path.includes('review')) score += 20;
    if (path.includes('pagination')) score += 15;
    if (path.includes('page')) score += 10;
    if (cls.includes('next')) score += 5;
    let tag = (el.tagName || '').toLowerCase();
    if (tag === 'button' || tag === 'a') score += 3;
    nextCandidates.push({ el, score, text, cls, tag: el.tagName });
}

nextCandidates.sort((a, b) => b.score - a.score);

if (nextCandidates.length > 0) {
    let best = nextCandidates[0];
    best.el.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true, view: window }));
    return { clicked: true, text: best.text, tag: best.tag, cls: best.cls, source: 'fallback-scan', score: best.score };
}

return { clicked: false, text: '', tag: '', cls: '', source: 'not-found' };
"""

# 리뷰 페이지 전환 검증용 - 현재 보이는 리뷰 내용의 핑거프린트를 반환
JS_GET_REVIEW_FINGERPRINT = """
try {
    let inProd = document.querySelector('oy-review-review-in-product');
    if (!inProd || !inProd.shadowRoot) return '__no_component__';
    let root1 = inProd.shadowRoot;
    let listProvider = root1.querySelector('oy-review-review-list-provider');
    if (!listProvider) return '__no_provider__';
    let reviewList = listProvider.querySelector('oy-review-review-list');
    if (!reviewList || !reviewList.shadowRoot) return '__no_list__';
    let items = reviewList.shadowRoot.querySelectorAll('oy-review-review-item');
    if (!items.length) return '__no_items__';
    let fp = '';
    for (let i = 0; i < Math.min(3, items.length); i++) {
        let item = items[i];
        if (!item.shadowRoot) continue;
        let contentEl = item.shadowRoot.querySelector('oy-review-review-content');
        if (!contentEl || !contentEl.shadowRoot) continue;
        let p = contentEl.shadowRoot.querySelector('.content p');
        if (p) fp += '|' + p.textContent.trim().substring(0, 40);
    }
    return fp || '__empty__';
} catch(e) { return '__error__:' + e.message; }
"""

def init_driver():
    """Selenium Edge 헤드리스 드라이버 초기화 (창 크기 1920x1080 강제 설정)"""
    print("\n[*] Selenium Edge 드라이버를 초기화하는 중입니다...")
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080") # 스크롤 차단 방지를 위한 넓은 뷰포트 강제 설정!
    options.binary_location = EDGE_BINARY_PATH
    # 자동화 차단 방지 헤더 설정
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    
    try:
        driver = webdriver.Edge(options=options)
        print("[*] Edge 드라이버 초기화 및 뷰포트 설정(1920x1080) 완료.")
        return driver
    except Exception as e:
        print(f"[!] 드라이버 초기화 실패: {e}")
        sys.exit(1)

def crawl_raw_reviews_from_page(driver, goods_no, page_info):
    """각 고유 상품 상세 페이지에서 페이지네이션을 차례대로 클릭하며 가능한 많은 로우(Raw) 리뷰 데이터를 수집"""
    url = f"https://www.oliveyoung.co.kr/store/goods/getGoodsDetail.do?goodsNo={goods_no}"
    print(f"\n" + "="*60)
    print(f"[*] 크롤링 타겟 페이지: [{goods_no}] {page_info['name']}")
    print(f"[*] URL: {url}")
    print(f"[*] 상세 페이지 이동 중...")
    
    driver.get(url)
    time.sleep(6) # 로딩 안정화 대기
    
    # 1. 리뷰 탭 클릭을 통해 리뷰 동적 컴포넌트 마운트
    print("[*] '리뷰&셔터' 탭 활성화 시도 중...")
    
    # 리뷰 위치로 먼저 스크롤을 살짝 내려서 버튼들이 마운트되도록 보장
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight * 0.3);")
    time.sleep(1.5)
    
    buttons = driver.find_elements(By.TAG_NAME, "button")
    clicked = False
    for btn in buttons:
        try:
            txt = btn.text
            if "리뷰&셔터" in txt or "리뷰 " in txt:
                driver.execute_script("arguments[0].click();", btn)
                clicked = True
                print(f"[*] 리뷰 탭 클릭 성공! (버튼 텍스트: '{txt}')")
                break
        except Exception:
            continue
            
    if not clicked:
        print("[!] 리뷰 탭 활성화 버튼을 직접 클릭하지 못했습니다. 디폴트로 열려있는지 확인합니다.")
        
    time.sleep(3.5)
    
    # ★ 핵심: 리뷰 탭 클릭 후 oy-review-pagination 컴포넌트가 완전히 마운트될 때까지
    # 페이지 하단 방향으로 충분히 스크롤하여 shadow DOM 렌더링을 강제로 유도
    print("[*] 리뷰 목록 및 페이지네이션 렌더링을 위해 스크롤 진행 중...")
    for _ in range(4):
        driver.execute_script("window.scrollBy(0, 600);")
        time.sleep(0.7)
    time.sleep(1.5)  # 페이지네이션 컴포넌트 마운트 안정화
    
    # 2. 페이지네이션 순회 수집 루프
    raw_reviews = {}
    max_pages = page_info["scroll_steps"]
    delay = page_info["delay"]
    
    print(f"[*] 페이지네이션 수집 루프 시작 (최대 {max_pages}페이지 진행)")
    
    for page in range(1, max_pages + 1):
        # 2-1. 매 페이지 수집 전, 리뷰 컨테이너 위치로 스크롤하여 리뷰 목록 + 페이지네이션 모두 뷰포트에 들어오도록
        try:
            driver.execute_script("""
                let el = document.querySelector('oy-review-review-in-product');
                if (el) {
                    el.scrollIntoView({behavior: 'instant', block: 'start'});
                } else {
                    window.scrollTo(0, document.body.scrollHeight * 0.5);
                }
            """)
            time.sleep(1.0)
            
            # 리뷰 목록 끝부분 + 페이지네이션이 뷰포트 안에 들어오도록 추가 스크롤
            driver.execute_script("window.scrollBy(0, 800);")
            time.sleep(1.5)
            
            # 클릭 전 현재 리뷰 핑거프린트 캡처 (페이지 전환 검증용)
            pre_fingerprint = driver.execute_script(JS_GET_REVIEW_FINGERPRINT)
            
            result = driver.execute_script(JS_PAGINATE_NEXT, page + 1)
            
            # JS가 dict를 반환: {clicked, text, tag, cls, source}
            if isinstance(result, dict):
                navigated = result.get('clicked', False)
                if navigated:
                    print(f"    [*] 클릭 → '{result.get('text')}' [{result.get('tag')}] cls={result.get('cls')[:40]} source={result.get('source')}")
            else:
                navigated = bool(result)
            
            if not navigated:
                # 1회 재시도
                print(f"    [~] Page {page+1} 버튼을 찾지 못함. 1.5초 후 재시도...")
                time.sleep(1.5)
                driver.execute_script("window.scrollBy(0, 400);")
                time.sleep(1.0)
                pre_fingerprint = driver.execute_script(JS_GET_REVIEW_FINGERPRINT)
                result = driver.execute_script(JS_PAGINATE_NEXT, page + 1)
                if isinstance(result, dict):
                    navigated = result.get('clicked', False)
                    if navigated:
                        print(f"    [*] 재시도 클릭 → '{result.get('text')}' [{result.get('tag')}] source={result.get('source')}")
                else:
                    navigated = bool(result)
            
        except Exception as e:
            print(f"    [!] 스크롤 포커싱 도중 예외 (계속 진행): {e}")

        # 2-2. 해당 페이지 리뷰 데이터 쉐도우 루트 파싱 JS 실행
        step_reviews = driver.execute_script(JS_DEEP_EXTRACT)
        if step_reviews:
            for r in step_reviews:
                key = (r['username'], r['date'], r['content'][:50])
                if key not in raw_reviews:
                    raw_reviews[key] = {
                        "goods_no": goods_no,
                        "username": r['username'],
                        "skin_types": r['skin_types'],
                        "rating": r['rating'],
                        "date": r['date'],
                        "content": r['content'],
                        "option_name": r['option_name']
                    }
                    
        print(f"    - [{page}/{max_pages}] 페이지 수집 완료 | 현재까지 누적 고유 리뷰 수: {len(raw_reviews)}개")
        
        # 2-3. 다음 페이지로 이동 클릭 시도 (JS_PAGINATE_NEXT 실행)
        try:
            # 페이지네이션이 완전히 렌더링될 수 있도록 충분히 스크롤 후 대기
            driver.execute_script("""
                let el = document.querySelector('oy-review-review-in-product');
                if (el) el.scrollIntoView({behavior: 'instant', block: 'end'});
            """)
            time.sleep(1.5)  # pagination 컴포넌트 렌더링 안정화 대기
            driver.execute_script("window.scrollBy(0, 300);")
            time.sleep(1.0)  # 추가 렌더링 대기
            
            # 클릭 전 현재 리뷰 핑거프린트 캡처 (페이지 전환 검증용)
            pre_fingerprint = driver.execute_script(JS_GET_REVIEW_FINGERPRINT)
            
            navigated = driver.execute_script(JS_PAGINATE_NEXT, page + 1)
            
            if not navigated:
                # 1회 재시도: 스크롤을 더 내린 뒤 다시 시도
                print(f"    [~] Page {page+1} 버튼을 찾지 못함. 1.5초 후 재시도...")
                time.sleep(1.5)
                driver.execute_script("window.scrollBy(0, 400);")
                time.sleep(1.0)
                pre_fingerprint = driver.execute_script(JS_GET_REVIEW_FINGERPRINT)
                navigated = driver.execute_script(JS_PAGINATE_NEXT, page + 1)
            
            if not navigated:
                print(f"    [!] 다음 페이지(Page {page+1})가 존재하지 않거나 더 이상 클릭할 버튼이 없어 순회를 조기 종료합니다.")
                break
            
            # ★ 핵심 검증: 클릭 후 리뷰 내용이 실제로 바뀌는지 확인 (최대 8초 대기)
            # 내용이 안 바뀌면 = 잘못된 버튼을 클릭한 것 → 즉시 감지하고 종료
            changed = False
            for wait_i in range(16):
                time.sleep(0.5)
                post_fingerprint = driver.execute_script(JS_GET_REVIEW_FINGERPRINT)
                if post_fingerprint != pre_fingerprint and not post_fingerprint.startswith('__'):
                    changed = True
                    break
            
            if not changed:
                print(f"    [!] 페이지 전환 후 리뷰 내용이 변경되지 않음 - 클릭 타겟이 잘못됐거나 마지막 페이지입니다. 수집 종료.")
                print(f"    [!] 현재 핑거프린트: {post_fingerprint[:80]}")
                break
                
        except Exception as e:
            print(f"    [!] 페이지네이션 클릭 시도 중 에러 발생 (수집 조기 종료): {e}")
            break
            
        # 페이지 전환 확인 후 추가 렌더링 안정화 대기
        time.sleep(max(delay, 1.0) + random.uniform(0.3, 0.7))
        
    print(f"[*] 수집 종료. [{goods_no}] 페이지에서 총 {len(raw_reviews)}개의 고유 리뷰를 획득했습니다.")
    return list(raw_reviews.values())


def map_reviews_to_target_pads(all_raw_reviews):
    """수집된 대량의 리뷰 풀에 대해 옵션명 및 상품명을 분석하여 스킨푸드 11종 패드로 분류"""
    print("\n" + "="*60)
    print("[*] 11종 스킨푸드 패드로 리뷰 매핑 및 분류를 시작합니다...")
    print("="*60)
    
    # 11종 각 패드별 리뷰 수집소 생성
    categorized_reviews = {pad_name: [] for pad_name in TARGET_PADS.keys()}
    unmapped_count = 0
    
    for r in all_raw_reviews:
        mapped = False
        opt = (r['option_name'] or "").lower()
        g_no = r['goods_no']
        
        # 1. 1차 매핑: 리뷰 카드 내부의 옵션명(option_name) 분석
        for pad_name, info in TARGET_PADS.items():
            for kw in info["keywords"]:
                if kw in opt:
                    categorized_reviews[pad_name].append(r)
                    mapped = True
                    break
            if mapped:
                break
                
        # 2. 2차 매핑 (단품 전용 페이지 대응): 
        # 옵션명이 비어있거나 매핑에 실패한 경우, 해당 리뷰를 긁어온 상품 코드의 디폴트 타겟 패드로 분류
        if not mapped:
            for pad_name, info in TARGET_PADS.items():
                if g_no == info["default_goods"]:
                    categorized_reviews[pad_name].append(r)
                    mapped = True
                    break
                
        if not mapped:
            unmapped_count += 1
            
    print("[*] 매핑 결과 분류 통계:")
    for pad, revs in categorized_reviews.items():
        print(f"    - {pad}: {len(revs)}개 분류 완료")
    print(f"    - 분류되지 않은 기타 옵션 리뷰: {unmapped_count}개")
    
    return categorized_reviews

def balance_reviews(pad_name, reviews, target_total=50):
    """
    평점 균등 분배 샘플링 알고리즘 (Recursive Rating Balancing Algorithm)
    1~5점의 별점을 최대한 10개씩 골고루 섞되, 특정 저평가 리뷰가 부족할 경우 
    남는 목표 수량을 다른 평점 그룹에 균등하게 재귀 재배분하여 정확히 50개를 채움.
    """
    print(f"\n[*] [{pad_name}] 평점 균등 샘플링 진행 중...")
    
    # 평점 그룹별(1점~5점) 리뷰 분류
    reviews_by_rating = {1: [], 2: [], 3: [], 4: [], 5: []}
    for r in reviews:
        score = r["rating"]
        if score in reviews_by_rating:
            reviews_by_rating[score].append(r)
            
    ratings = [1, 2, 3, 4, 5]
    available = {r: len(reviews_by_rating[r]) for r in ratings}
    allocated = {r: 0 for r in ratings}
    
    total_available = sum(available.values())
    if total_available <= target_total:
        print(f"    [!] 가용 리뷰 수({total_available})가 목표량({target_total}) 이하입니다. 전체를 선택합니다.")
        return reviews
        
    remaining_target = target_total
    active_ratings = list(ratings)
    
    # 재귀 배분 루프
    while remaining_target > 0 and active_ratings:
        num_groups = len(active_ratings)
        base_share = remaining_target // num_groups
        if base_share == 0:
            base_share = 1
            
        next_active = []
        for r in active_ratings:
            needed = base_share
            av = available[r] - allocated[r]
            
            if av <= needed:
                allocated[r] += av
                remaining_target -= av
            else:
                allocated[r] += needed
                remaining_target -= needed
                next_active.append(r)
                
        # 한 라운드를 마친 뒤 분배가 교착상태인 경우 비례 추가 분배
        if len(next_active) == len(active_ratings) and remaining_target > 0:
            for r in sorted(next_active, key=lambda x: available[x] - allocated[x], reverse=True):
                if remaining_target > 0 and (available[r] - allocated[r]) > 0:
                    allocated[r] += 1
                    remaining_target -= 1
            break
            
        active_ratings = next_active
        
    # 최종 선택 및 셔플링
    selected_reviews = []
    for r in ratings:
        count = allocated[r]
        revs = reviews_by_rating[r]
        random.shuffle(revs)
        selected_reviews.extend(revs[:count])
        print(f"    - 평점 {r}점: 총 {available[r]}개 중 {count}개 선택")
        
    print(f"    -> 최종 선정 완료: {len(selected_reviews)}개")
    return selected_reviews

def save_to_premium_excel(final_reviews_dict):
    """최종 선별된 11종 패드의 균등 평점 리뷰 데이터를 프리미엄 디자인이 적용된 단일 XLSX 시트로 저장"""
    print(f"\n[*] 엑셀 파일 생성을 준비 중입니다... (파일명: {OUTPUT_FILENAME})")
    
    flat_data = []
    for pad_name, revs in final_reviews_dict.items():
        for r in revs:
            flat_data.append({
                "타겟상품명": pad_name,
                "올리브영 상품코드": r["goods_no"],
                "구매 옵션명": r["option_name"] if r["option_name"] else "단품",
                "작성자": r["username"],
                "피부타입": r["skin_types"],
                "별점": f"{r['rating']}점",
                "작성일": r["date"],
                "리뷰 내용": r["content"]
            })
            
    df = pd.DataFrame(flat_data)
    if df.empty:
        print("[!] 최종 수집 데이터가 존재하지 않아 엑셀 파일을 생성할 수 없습니다.")
        return
        
    columns_order = ["타겟상품명", "올리브영 상품코드", "구매 옵션명", "작성자", "피부타입", "별점", "작성일", "리뷰 내용"]
    df = df[columns_order]
    
    try:
        with pd.ExcelWriter(OUTPUT_FILENAME, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='스킨푸드 패드 리뷰')
            
            workbook = writer.book
            worksheet = writer.sheets['스킨푸드 패드 리뷰']
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # 프리미엄 뷰티 올리브 그린 칼라 팔레트 디자인 적용
            header_fill = PatternFill(start_color="33691E", end_color="33691E", fill_type="solid") # 깊은 올리브 그린 헤더
            header_font = Font(name="Malgun Gothic", size=11, bold=True, color="FFFFFF")
            data_font = Font(name="Malgun Gothic", size=10)
            
            thin_border = Border(
                left=Side(style='thin', color='EAEAEA'),
                right=Side(style='thin', color='EAEAEA'),
                top=Side(style='thin', color='EAEAEA'),
                bottom=Side(style='thin', color='EAEAEA')
            )
            
            # 헤더 디자인 반영
            for col_idx in range(1, len(columns_order) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
            # 데이터 로우 스타일링 및 정렬 처리
            for row in range(2, len(df) + 2):
                for col in range(1, len(columns_order) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = thin_border
                    
                    # 칼럼별 정렬 최적화
                    # 타겟상품명, 상품코드, 작성자, 별점, 작성일은 정가운데 정렬하여 시각화 효과 극대화
                    if col in [1, 2, 4, 6, 7]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    # 리뷰 내용은 왼쪽 정렬 및 자동 개행 (텍스트 랩핑)
                    elif col == 8:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        
            # 행의 높이 증대
            worksheet.row_dimensions[1].height = 32
            for r in range(2, len(df) + 2):
                worksheet.row_dimensions[r].height = 25
                
            # 열 너비 자동 최적화 맞춤
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val = str(cell.value or '')
                    # 한글 보정 글자 수 계산
                    byte_len = len(val.encode('utf-8'))
                    visual_len = (byte_len + len(val)) // 2
                    if visual_len > max_len:
                        max_len = visual_len
                # 너무 과도하게 긴 칼럼(리뷰 내용 등) 방지를 위한 적정 한도 제어
                worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 70)
                
        print(f"\n[+] 프리미엄 엑셀 파일 저장 성공: {os.path.abspath(OUTPUT_FILENAME)}")
        print(f"[+] 최종 저장된 데이터 수: {len(df)}개 행 (11개 제품군 x 최대 50개씩)")
    except Exception as e:
        print(f"[!] 엑셀 저장 도중 에러가 발생했습니다: {e}")

def main():
    print("="*60)
    print("      올리브영 스킨푸드 11종 패드 균등 평점 리뷰 수집기")
    print("="*60)
    
    # 1. Selenium 헤드리스 Edge 드라이버 구동
    driver = init_driver()
    
    all_pages_reviews = []
    
    try:
        # 2. 고유 상품 페이지 6개 목록을 돌며 각각 Raw 리뷰 덩어리 수집
        for idx, (goods_no, page_info) in enumerate(PRODUCT_PAGES.items()):
            print(f"\n[*] [페이지 진행도: {idx+1}/{len(PRODUCT_PAGES)}]")
            page_reviews = crawl_raw_reviews_from_page(driver, goods_no, page_info)
            all_pages_reviews.extend(page_reviews)
            
            # 서버 부하 조절용 세션 전환 쿨타임 대기
            if idx < len(PRODUCT_PAGES) - 1:
                print("[*] 다음 상품 페이지로 이동하기 전 4초간 쿨다운 대기합니다...")
                time.sleep(4)
                
    finally:
        print("\n[*] Selenium 드라이버를 정상 종료합니다.")
        driver.quit()
        
    # 3. 수집된 전체 로우 리뷰를 스킨푸드 11종 패드로 라우팅 분류
    categorized_reviews = map_reviews_to_target_pads(all_pages_reviews)
    
    # 4. 각 패드별 평점 균등 배분 샘플링 실행 (목표 2000건 달성을 위해 패드별 최대 185건씩 선별)
    final_selected_reviews = {}
    for pad_name, rev_pool in categorized_reviews.items():
        balanced = balance_reviews(pad_name, rev_pool, target_total=185)
        final_selected_reviews[pad_name] = balanced
        
    # 5. 스타일링된 완성도 높은 프리미엄 엑셀 파일로 출력
    save_to_premium_excel(final_selected_reviews)
    
    print("\n" + "="*60)
    print("      모든 크롤링 및 균등 평점 리뷰 데이터 수집 작업 완료!")
    print("="*60)

if __name__ == "__main__":
    main()
